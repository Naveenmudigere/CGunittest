import openpyxl


def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)

def ReadData(file,sheetName,rownum,columno):
    workbook = openpyxl.load_workbook(file)
    sheet =workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=columno).value


def WriteData(file,sheetName,rownum,columnno,data):
    workbook =openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum,column=columnno).value=data
    workbook.save(file)


